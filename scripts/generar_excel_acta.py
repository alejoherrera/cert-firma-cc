"""Genera el acta tabular en Excel (.xlsx) a partir de listado_hashes.json.

Uso:
    python scripts/generar_excel_acta.py [lote_dir]

Default: el lote mas reciente bajo data/output/lote_*/

Output: <lote_dir>/acta.xlsx

Formato amigable para revision administrativa, filtrado, copia a portapapeles.
La firma digital institucional va sobre el acta.pdf (no sobre el xlsx).
"""
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_BASE = REPO_ROOT / 'data' / 'output'

AZUL_CGR_HEX = '222D72'
GRIS_HEADER_HEX = 'EAEAEA'
GRIS_ALT_HEX = 'F6F8FB'


def safe(s: str) -> str:
    return s.encode('ascii', errors='replace').decode('ascii')


def ultimo_lote() -> Path:
    lotes = sorted(OUTPUT_BASE.glob('lote_*'), reverse=True)
    if not lotes:
        sys.exit(f"[ERROR] No hay lotes en {OUTPUT_BASE}")
    return lotes[0]


def construir_xlsx(lote_dir: Path) -> Path:
    listado_path = lote_dir / 'listado_hashes.json'
    manifest_path = lote_dir / 'manifest.json'
    if not listado_path.exists() or not manifest_path.exists():
        sys.exit(f"[ERROR] Lote incompleto: faltan archivos en {lote_dir}")

    registros = json.loads(listado_path.read_text(encoding='utf-8'))
    manifest = json.loads(manifest_path.read_text(encoding='utf-8'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Acta'

    azul_fill = PatternFill(start_color=AZUL_CGR_HEX, end_color=AZUL_CGR_HEX, fill_type='solid')
    gris_fill = PatternFill(start_color=GRIS_HEADER_HEX, end_color=GRIS_HEADER_HEX, fill_type='solid')
    alt_fill = PatternFill(start_color=GRIS_ALT_HEX, end_color=GRIS_ALT_HEX, fill_type='solid')

    font_titulo = Font(name='Calibri', size=14, bold=True, color=AZUL_CGR_HEX)
    font_subt = Font(name='Calibri', size=11, color=AZUL_CGR_HEX)
    font_meta_label = Font(name='Calibri', size=10, bold=True, color=AZUL_CGR_HEX)
    font_meta_val = Font(name='Calibri', size=10)
    font_meta_hash = Font(name='Consolas', size=9)
    font_header_tabla = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Calibri', size=10.5)
    font_body_hash = Font(name='Consolas', size=9.5)
    font_pie = Font(name='Calibri', size=10, italic=True, color='4A4A4A')

    thin = Side(border_style='thin', color='CCCCCC')
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)

    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    izq = Alignment(horizontal='left', vertical='center', wrap_text=True)
    derecha = Alignment(horizontal='right', vertical='center')

    # Anchos de columna
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 70

    fila = 1

    # Titulo institucional
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value='CONTRALORÍA GENERAL DE LA REPÚBLICA')
    c.font = font_titulo
    c.alignment = centro
    fila += 1

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value='División de Gestión de Apoyo · Unidad Centro de Capacitación')
    c.font = font_subt
    c.alignment = centro
    fila += 2

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value='ACTA DE EMISIÓN DE CERTIFICADOS')
    c.font = Font(name='Calibri', size=13, bold=True)
    c.alignment = centro
    fila += 2

    # Metadata
    fecha_emision_lote = registros[0]['fecha_emision'] if registros else '—'
    meta = [
        ('Curso / Actividad', manifest.get('curso') or '(sin especificar)'),
        ('Fecha de emisión de los certificados', fecha_emision_lote),
        ('Total de certificados emitidos', str(manifest['total_certificados'])),
        ('Versión del hash canónico', f"v{manifest['hash_version']}"),
        ('Identificador del lote', manifest['lote_id']),
        ('Fecha de generación del acta', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Git commit', manifest.get('git_commit', '')),
        ('SHA-256 del listado de hashes', manifest['hash_listado_json_sha256']),
    ]
    for label, val in meta:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        ws.merge_cells(start_row=fila, start_column=3, end_row=fila, end_column=4)
        cl = ws.cell(row=fila, column=1, value=label)
        cl.font = font_meta_label
        cl.fill = gris_fill
        cl.alignment = izq
        cl.border = border_all
        cv = ws.cell(row=fila, column=3, value=val)
        cv.font = font_meta_hash if label.startswith('SHA-256') else font_meta_val
        cv.alignment = izq
        cv.border = border_all
        fila += 1

    fila += 1

    # Prosa institucional
    prosa = (
        "Quien suscribe, en calidad de Jefatura A.I de la Unidad Centro de Capacitación de la "
        "División de Gestión de Apoyo de la Contraloría General de la República, hace constar la "
        "emisión de los certificados de participación cuyos hashes criptográficos SHA-256 se listan "
        "a continuación. Cada certificado emitido contiene impreso su propio hash, el cual debe "
        "coincidir con la entrada correspondiente en este acta. La firma digital aplicada sobre el "
        "acta PDF da fe institucional del lote completo aquí registrado."
    )
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value=prosa)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.font = Font(name='Calibri', size=10)
    ws.row_dimensions[fila].height = 60
    fila += 2

    # Tabla principal: encabezado
    headers = ['Nº', 'Participante', 'Fecha emisión', 'Hash SHA-256']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=fila, column=col, value=h)
        c.font = font_header_tabla
        c.fill = azul_fill
        c.alignment = centro
        c.border = border_all
    fila += 1
    fila_inicio_tabla = fila

    # Tabla principal: cuerpo
    for i, r in enumerate(registros):
        c1 = ws.cell(row=fila, column=1, value=r['nro'])
        c1.font = font_body
        c1.alignment = centro

        c2 = ws.cell(row=fila, column=2, value=r['nombre'])
        c2.font = font_body
        c2.alignment = izq

        c3 = ws.cell(row=fila, column=3, value=r['fecha_emision'])
        c3.font = font_body
        c3.alignment = izq

        c4 = ws.cell(row=fila, column=4, value=r['hash_sha256'])
        c4.font = font_body_hash
        c4.alignment = izq

        for col in (1, 2, 3, 4):
            ws.cell(row=fila, column=col).border = border_all
            if i % 2 == 1:
                ws.cell(row=fila, column=col).fill = alt_fill
        fila += 1

    fila_fin_tabla = fila - 1

    # Habilitar AutoFilter sobre la tabla
    rango = (
        f"A{fila_inicio_tabla - 1}:"
        f"D{fila_fin_tabla}"
    )
    ws.auto_filter.ref = rango
    # Freeze panes: congelar encabezado de tabla
    ws.freeze_panes = ws.cell(row=fila_inicio_tabla, column=1)

    fila += 1

    # Pie firmante
    pie_lineas = [
        ('El presente documento será firmado digitalmente por:', True),
        ('Lic. Juan Alejandro Herrera López', False),
        ('Jefatura A.I, Unidad Centro de Capacitación', False),
        ('División de Gestión de Apoyo', False),
        ('Contraloría General de la República', False),
    ]
    for txt, bold in pie_lineas:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
        c = ws.cell(row=fila, column=1, value=txt)
        if bold:
            c.font = Font(name='Calibri', size=10.5, bold=True)
        else:
            c.font = font_pie
        c.alignment = centro
        fila += 1

    out = lote_dir / 'acta.xlsx'
    wb.save(out)
    return out


def main():
    if len(sys.argv) > 1:
        lote = Path(sys.argv[1])
    else:
        lote = ultimo_lote()
    if not lote.exists() or not lote.is_dir():
        sys.exit(f"[ERROR] No existe lote: {lote}")
    print(f"[INFO] Lote: {safe(lote.name)}")
    out = construir_xlsx(lote)
    print(f"[OK] Acta xlsx -> {safe(str(out))} ({out.stat().st_size:,} bytes)")


if __name__ == '__main__':
    main()
